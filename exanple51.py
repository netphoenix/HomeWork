from openpyxl import Workbook

# создание нового файла
wb = Workbook()

# выбор активного листа
ws = wb.active

# запись данных в ячейки
ws['A1'] = 'Hello'
ws['B1'] = 'World!'

# создание нового листа и выбор его
ws2 = wb.create_sheet("Sheet2")
wb.active = 1

# запись данных в новый лист
ws2['A1'] = 'This'
ws2['B1'] = 'is'
ws2['C1'] = 'Sheet2'

# сохранение файла
wb.save("example.xlsx")
